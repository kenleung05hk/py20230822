import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# Set the directory path where the xlsx files are located
directory_path = 'C:\\Users\\KohleKen\\OneDrive - Kohle Services Limited\\桌面\\出金報表'

# Find all files that start with 'KH2024'
kh2024_files = []
for filename in os.listdir(directory_path):
    if filename.endswith('.xlsx') and filename.startswith('KH2024'):
        kh2024_files.append(filename)
print(kh2024_files)

#Reordered list UT to last
my_list = kh2024_files
# Define the order of the list
my_order = [i for i in range(len(my_list)) if not my_list[i].endswith('UT.xlsx')] + [i for i in range(len(my_list)) if my_list[i].endswith('UT.xlsx')]
# Reorder the list
kh2024_files = [my_list[i] for i in my_order]



'''
for file in kh2024_files:
    wb = load_workbook(os.path.join(directory_path, file))
# set column 5 of default sheet to be text:
    ws = wb.active
    for row in ws[2:ws.max_row]: # skip the header
        cell = row[4] # column 5
        cell.number_format = numbers.FORMAT_TEXT
        print(cell.value)
    wb.save(os.path.join(directory_path, file))
'''



# Loop through all the files
df_list = []
df_list_fail = []
for file in kh2024_files:
    # Read the file
    df = pd.read_excel(os.path.join(directory_path, file), skiprows=7,dtype={4:str})
    # Check if the column '下发状态' or '下發狀態' exists in the dataframe
    if '下发状态' in df.columns:
        # Filter the rows where column '下发状态' is "成功"
        df = df[df['下发状态'] == '成功']
    elif '下發狀態' in df.columns:
        # Filter the rows where column '下發狀態' is "成功"
        df = df[df['下發狀態'] == '成功']
    else:
        print("Column '下发状态' or '下發狀態' does not exist in the dataframe.")
        continue
    df_list.append(df)

    sum_of_df = df[df.columns[3]].sum()
    print(f"listsuccess {file} : {sum_of_df}.")
    

for file in kh2024_files:
    # Read the file
    df = pd.read_excel(os.path.join(directory_path, file), skiprows=7,dtype={4:str})
    # Check if the column '下发状态' or '下發狀態' exists in the dataframe
    if '下发状态' in df.columns:
        # Filter the rows where column '下发状态' is "成功"
        df = df[df['下发状态'] != '成功']
    elif '下發狀態' in df.columns:
        # Filter the rows where column '下發狀態' is "成功"
        df = df[df['下發狀態'] != '成功']
    else:
        print("Column '下发状态' or '下發狀態' does not exist in the dataframe.")
        continue
    df_list_fail.append(df)

    sum_of_df = df[df.columns[3]].sum()
    print(f"listfail {file} : {sum_of_df}.")


# Concatenate all dataframes into one
merged_df = pd.concat(df_list)
# Save the merged dataframe to a new file
merged_df.to_excel(os.path.join(directory_path, 'merged_data_已完成.xlsx'), index=False)

# Concatenate all dataframes into one
merged_df1 = pd.concat(df_list_fail)
# Save the merged dataframe to a new file
merged_df1.to_excel(os.path.join(directory_path, 'merged_data_已失敗.xlsx'), index=False)





# write the DataFrame to an Excel file
#wb = Workbook()
#ws = wb.active
#for r in dataframe_to_rows(merged_df, index=False, header=True):
    #ws.append(r)
    #print(r)

# set the format of column 4 to text
#for cell in ws['E']:
    #cell.number_format = '0'

# save the Excel file
#wb.save(os.path.join(directory_path, 'merged_data_已完成.xlsx'))
input()
